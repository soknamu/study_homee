from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 엑셀 워크북 생성
workbook = Workbook()
sheet = workbook.active

# 데이터 입력
data = [
    [1, 4, 15, 25, 35],
    [2, 4, 15, 25, 35],
    [3, 4, 15, 25, 35],
    [4, 4, 15, 25, 35],
    [5, 4, 15, 25, 35],
    [6, 4, 15, 25, 35],
    [7, 4, 15, 25, 35],
    [8, 4, 15, 25, 35]
]

for row in data:
    sheet.append(row)

# 유통기한 계산 수식 적용
for row in range(2, sheet.max_row + 1):
    min_cfu_range = [f"{get_column_letter(col)}{row}" for col in range(2, sheet.max_column)]
    formula = f"=MAX(0, MIN({','.join(min_cfu_range)}))"
    sheet.cell(row=row, column=sheet.max_column).value = formula
    sheet.cell(row=row, column=sheet.max_column).alignment = Alignment(horizontal='right')

# 엑셀 파일 저장
workbook.save('./excel/유통기한_계산.xlsx')